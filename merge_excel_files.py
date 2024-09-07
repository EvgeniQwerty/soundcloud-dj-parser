import os
import openpyxl

# Function to read data from an Excel file
def read_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []

    # Read each row from the Excel file and append to the data list
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        data.append(row)

    return data

# Function to write data to an Excel file
def write_data_to_excel(data, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the header
    sheet.append(["User URL", "User Name", "Email"])

    # Write the data
    for row in data:
        sheet.append(row)

    # Save the Excel file
    workbook.save(output_file)

# Function to merge all Excel files in the current directory
def merge_excel_files(output_file="dj_emails.xlsx"):
    # Get a list of all Excel files in the current directory, excluding the output file
    excel_files = [f for f in os.listdir() if f.endswith('.xlsx') and f != output_file]

    if not excel_files:
        print("No Excel files found in the current directory.")
        return

    merged_data = []
    seen_entries = set()  # To keep track of unique entries

    # Iterate over each Excel file
    for file in excel_files:
        file_data = read_data_from_excel(file)

        # Iterate over each row of the data
        for row in file_data:
            # Check if the entry (User URL, Email) is unique
            user_url = row[0]
            email = row[2]
            if (user_url, email) not in seen_entries:
                merged_data.append(row)
                seen_entries.add((user_url, email))

    # Write the merged data to the output Excel file
    write_data_to_excel(merged_data, output_file)
    print(f"Merged data saved to {output_file}.")

    # Remove individual Excel files after merging
    for file in excel_files:
        os.remove(file)
        print(f"Deleted {file}")

# Call the function
if __name__ == "__main__":
    merge_excel_files()